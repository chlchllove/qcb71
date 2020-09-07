'''
接口测试自动化的步骤：
1、接口测试用例 --- Done
2、Python代码读取接口测试用例 --- Done--read_date()
3、requests 库发送接口请求 --- Done=== api_request
4、执行结果 vs  预期结果  == 用例执行是否是通过的！--结果
5、结果回写到excel里  -- openpyxl --Done -- write_result()
函数定义：
1、实现功能  2、 参数 --变化的值  3、 返回值--- 别人需要从你这里得到的数据
'''
import openpyxl
import requests
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    case_list = []
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i, column=1).value,
        url = sheet.cell(row=i, column=5).value,
        data = sheet.cell(row=i, column=6).value,
        expected = sheet.cell(row=i, column=7).value
        )
        case_list.append(case)
    return case_list


#发送接口请求
def request(api_url,api_data):
    qcd_header ={"X-Lemonban-Media-Type":"lemonban.v2", "Content-Type":"application/json"}
    response = requests.post(url=api_url, json=api_data, headers=qcd_header)
    return response.json()

def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)

def execute_func(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")
        data = case["data"]
        data = eval(data)
        expected = case.get("expected")
        expected = eval(expected)
        real_result = request(api_url=url,api_data=data)
        real_msg = real_result["msg"]
        expected_msg = expected.get("msg")
        print("执行结果是：{}".format(real_msg))
        print("执行结果是：{}".format(expected_msg))
        if real_msg == expected_msg:
            print("第{}条执行通过".format(case_id))
            final_result = "PASSED"
        else:
            print("第{}条执行不通过".format(case_id))
            final_result = "FAILED"
        print("*" * 20)
        write_result(filename,sheetname,case_id+1,8,final_result)
execute_func("test_case_api.xlsx","register")